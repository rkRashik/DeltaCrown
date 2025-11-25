"""
Response Export Service

Export registration form responses to various formats (CSV, Excel, JSON).
Includes filtering, analytics, and bulk operations.
"""
import csv
import json
from io import StringIO, BytesIO
from typing import List, Dict, Optional
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Q
from apps.tournaments.models import FormResponse, TournamentRegistrationForm


class ResponseExportService:
    """
    Service for exporting form responses to various formats.
    
    Supports:
    - CSV export with custom field selection
    - Excel export with formatting
    - JSON export with full data
    - Filtered exports by status, date range, payment status
    - Analytics summary inclusion
    """
    
    def __init__(self, tournament_form_id: int):
        self.tournament_form = TournamentRegistrationForm.objects.select_related(
            'template', 'tournament'
        ).get(id=tournament_form_id)
        self.form_schema = self.tournament_form.template.form_schema
    
    def _get_all_field_ids(self) -> List[Dict]:
        """Extract all field IDs and labels from form schema"""
        fields = []
        for section in self.form_schema.get('sections', []):
            for field in section.get('fields', []):
                # Skip non-data fields
                if field.get('type') not in ['section_header', 'divider']:
                    fields.append({
                        'id': field.get('id'),
                        'label': field.get('label', field.get('id')),
                        'type': field.get('type'),
                        'section': section.get('title', 'General'),
                    })
        return fields
    
    def _apply_filters(
        self,
        queryset,
        status: Optional[List[str]] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        has_paid: Optional[bool] = None,
        payment_verified: Optional[bool] = None,
        search: Optional[str] = None,
    ):
        """Apply filters to response queryset"""
        
        if status:
            queryset = queryset.filter(status__in=status)
        
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        if has_paid is not None:
            queryset = queryset.filter(has_paid=has_paid)
        
        if payment_verified is not None:
            queryset = queryset.filter(payment_verified=payment_verified)
        
        if search:
            # Search in response data (JSON field search)
            queryset = queryset.filter(
                Q(user__username__icontains=search) |
                Q(team__name__icontains=search)
            )
        
        return queryset
    
    def export_to_csv(
        self,
        status: Optional[List[str]] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        has_paid: Optional[bool] = None,
        payment_verified: Optional[bool] = None,
        search: Optional[str] = None,
        include_metadata: bool = True,
    ) -> HttpResponse:
        """
        Export responses to CSV format.
        
        Returns HttpResponse with CSV file ready for download.
        """
        # Get filtered responses
        responses = FormResponse.objects.filter(
            tournament_form=self.tournament_form
        ).select_related('user', 'team')
        
        responses = self._apply_filters(
            responses, status, date_from, date_to,
            has_paid, payment_verified, search
        )
        
        # Create CSV
        output = StringIO()
        
        # Get all fields
        all_fields = self._get_all_field_ids()
        
        # Define CSV headers
        headers = ['ID', 'Status', 'Submitted At']
        
        if include_metadata:
            headers.extend([
                'User', 'Team', 'Has Paid', 'Payment Verified',
                'Created At', 'Updated At'
            ])
        
        # Add field headers
        headers.extend([f['label'] for f in all_fields])
        
        writer = csv.writer(output)
        writer.writerow(headers)
        
        # Write data rows
        for response in responses:
            row = [
                response.id,
                response.get_status_display(),
                response.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if response.submitted_at else 'N/A',
            ]
            
            if include_metadata:
                row.extend([
                    response.user.username if response.user else 'N/A',
                    response.team.name if response.team else 'N/A',
                    'Yes' if response.has_paid else 'No',
                    'Yes' if response.payment_verified else 'No',
                    response.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    response.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                ])
            
            # Add field values
            for field in all_fields:
                value = response.response_data.get(field['id'], '')
                
                # Format value based on type
                if isinstance(value, list):
                    value = ', '.join(str(v) for v in value)
                elif isinstance(value, dict):
                    value = json.dumps(value)
                
                row.append(str(value) if value else '')
            
            writer.writerow(row)
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv'
        )
        
        filename = f"{self.tournament_form.tournament.slug}_registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def export_to_excel(
        self,
        status: Optional[List[str]] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        has_paid: Optional[bool] = None,
        payment_verified: Optional[bool] = None,
        search: Optional[str] = None,
    ) -> HttpResponse:
        """
        Export responses to Excel format with formatting.
        
        Requires openpyxl package.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        # Get filtered responses
        responses = FormResponse.objects.filter(
            tournament_form=self.tournament_form
        ).select_related('user', 'team')
        
        responses = self._apply_filters(
            responses, status, date_from, date_to,
            has_paid, payment_verified, search
        )
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Registrations"
        
        # Get all fields
        all_fields = self._get_all_field_ids()
        
        # Define headers
        headers = [
            'ID', 'Status', 'User', 'Team',
            'Submitted At', 'Has Paid', 'Payment Verified'
        ]
        headers.extend([f['label'] for f in all_fields])
        
        # Write headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data rows
        for row_num, response in enumerate(responses, 2):
            ws.cell(row=row_num, column=1, value=response.id)
            ws.cell(row=row_num, column=2, value=response.get_status_display())
            ws.cell(row=row_num, column=3, value=response.user.username if response.user else 'N/A')
            ws.cell(row=row_num, column=4, value=response.team.name if response.team else 'N/A')
            ws.cell(row=row_num, column=5, value=response.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if response.submitted_at else 'N/A')
            ws.cell(row=row_num, column=6, value='Yes' if response.has_paid else 'No')
            ws.cell(row=row_num, column=7, value='Yes' if response.payment_verified else 'No')
            
            # Add field values
            for col_num, field in enumerate(all_fields, 8):
                value = response.response_data.get(field['id'], '')
                
                # Format value
                if isinstance(value, list):
                    value = ', '.join(str(v) for v in value)
                elif isinstance(value, dict):
                    value = json.dumps(value)
                
                ws.cell(row=row_num, column=col_num, value=str(value) if value else '')
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.cell(row=1, column=1, value="Registration Summary").font = Font(bold=True, size=14)
        
        summary_ws.cell(row=3, column=1, value="Tournament:")
        summary_ws.cell(row=3, column=2, value=self.tournament_form.tournament.name)
        
        summary_ws.cell(row=4, column=1, value="Template:")
        summary_ws.cell(row=4, column=2, value=self.tournament_form.template.name)
        
        summary_ws.cell(row=5, column=1, value="Total Responses:")
        summary_ws.cell(row=5, column=2, value=responses.count())
        
        summary_ws.cell(row=6, column=1, value="Export Date:")
        summary_ws.cell(row=6, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Status breakdown
        summary_ws.cell(row=8, column=1, value="Status Breakdown:").font = Font(bold=True)
        status_counts = {}
        for response in responses:
            status_counts[response.get_status_display()] = status_counts.get(response.get_status_display(), 0) + 1
        
        for idx, (status, count) in enumerate(status_counts.items(), 9):
            summary_ws.cell(row=idx, column=1, value=status)
            summary_ws.cell(row=idx, column=2, value=count)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"{self.tournament_form.tournament.slug}_registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def export_to_json(
        self,
        status: Optional[List[str]] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        has_paid: Optional[bool] = None,
        payment_verified: Optional[bool] = None,
        search: Optional[str] = None,
        include_analytics: bool = False,
    ) -> HttpResponse:
        """
        Export responses to JSON format with full data.
        
        Optionally includes analytics summary.
        """
        # Get filtered responses
        responses = FormResponse.objects.filter(
            tournament_form=self.tournament_form
        ).select_related('user', 'team')
        
        responses = self._apply_filters(
            responses, status, date_from, date_to,
            has_paid, payment_verified, search
        )
        
        # Build JSON structure
        data = {
            'tournament': {
                'id': self.tournament_form.tournament.id,
                'name': self.tournament_form.tournament.name,
                'slug': self.tournament_form.tournament.slug,
            },
            'template': {
                'id': self.tournament_form.template.id,
                'name': self.tournament_form.template.name,
            },
            'export_date': datetime.now().isoformat(),
            'total_count': responses.count(),
            'responses': [],
        }
        
        # Add analytics if requested
        if include_analytics:
            from apps.tournaments.services.form_analytics import FormAnalyticsService
            analytics = FormAnalyticsService(self.tournament_form.id)
            data['analytics'] = analytics.get_overview_metrics()
        
        # Add response data
        for response in responses:
            response_data = {
                'id': response.id,
                'status': response.status,
                'user': response.user.username if response.user else None,
                'team': response.team.name if response.team else None,
                'submitted_at': response.submitted_at.isoformat() if response.submitted_at else None,
                'has_paid': response.has_paid,
                'payment_verified': response.payment_verified,
                'payment_amount': float(response.payment_amount) if response.payment_amount else None,
                'created_at': response.created_at.isoformat(),
                'updated_at': response.updated_at.isoformat(),
                'form_data': response.response_data,
                'metadata': response.metadata,
            }
            data['responses'].append(response_data)
        
        # Create HTTP response
        response = HttpResponse(
            json.dumps(data, indent=2),
            content_type='application/json'
        )
        
        filename = f"{self.tournament_form.tournament.slug}_registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def get_export_preview(
        self,
        limit: int = 10,
        status: Optional[List[str]] = None,
    ) -> Dict:
        """
        Get preview of export data for UI display.
        
        Shows first N rows that will be exported.
        """
        responses = FormResponse.objects.filter(
            tournament_form=self.tournament_form
        ).select_related('user', 'team')
        
        if status:
            responses = responses.filter(status__in=status)
        
        responses = responses[:limit]
        
        all_fields = self._get_all_field_ids()
        
        preview_data = {
            'total_count': FormResponse.objects.filter(
                tournament_form=self.tournament_form,
                status__in=status if status else ['draft', 'submitted', 'approved', 'rejected']
            ).count(),
            'preview_count': responses.count(),
            'fields': all_fields,
            'rows': [],
        }
        
        for response in responses:
            row = {
                'id': response.id,
                'status': response.get_status_display(),
                'user': response.user.username if response.user else 'N/A',
                'submitted_at': response.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if response.submitted_at else 'N/A',
                'fields': {},
            }
            
            for field in all_fields:
                value = response.response_data.get(field['id'], '')
                if isinstance(value, list):
                    value = ', '.join(str(v) for v in value)
                row['fields'][field['id']] = str(value) if value else ''
            
            preview_data['rows'].append(row)
        
        return preview_data
